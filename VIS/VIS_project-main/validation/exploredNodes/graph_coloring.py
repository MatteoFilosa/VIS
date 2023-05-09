import json
from openpyxl import Workbook
from openpyxl import load_workbook

#workbook = Workbook()
wb_total = load_workbook(filename = 'total_count.xlsx')
sheet_total = wb_total.active

f = open('statechart.json')



#viridis

def color_assignment(number): #MAX : 743, INTERVALS: 20, 37
    
    if (number == 0):
        return "#fde725"
    if (0 < number <= 37):
        return "#dde318"
    if (37 < number <= 74):
        return "#bade28"
    if (74 < number <= 111):
        return "#95d840"
    if (111 < number <= 148):
        return "#75d054"
    if (148 < number <= 185):
        return "#56c667"
    if (185 < number <= 222):
        return "#3dbc74"
    if (222 < number <= 259):
        return "#29af7f"
    if (259 < number <= 296):
        return "#20a386"
    if (296 < number <= 333):
        return "#1f968b"
    if (333< number <= 370):
        return "#238a8d"
    if (370 < number <= 407):
        return "#287d8e"
    if (407 < number <= 444):
        return "#2d718e"
    if (444 < number <= 481):
        return "#33638d"
    if (481 < number <= 518):
        return "#39558c"
    if (518 < number <= 555):
        return "#404688" #############
    if (555 < number <= 592):
        return "#453781"
    if (592 < number <= 629):
        return "#33638d"
    if (629 < number <= 666):
        return "#482576"
    if (666 < number <= 702):
        return "#481467"
    if (number == 703):
        return "#440154"
    
iterator_marks = iter(f)

for el in f:
    
    
    print(el[0].name)
    
f.close()